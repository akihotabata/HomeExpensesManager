# -*- coding: utf-8 -*-
"""
家計簿GUI版 v0.7
PySide6 + SQLite / Excel版の基本機能をGUI化

主な機能
- 家計簿入力
- 家計簿一覧
- 全期間表示
- 月間表示：給与支給日基準
- 年間表示：1/1〜12/31
- 総収益 / 月間収益 / 年間収益
- Excel取込
- コンフィグ
  - 大区分設定
  - 中区分設定
  - 支払者設定
  - 給与支給日設定

集計ルール
- 支出 = 支出 + 金融貯蓄 + 現金/預金貯蓄
- 収益 = 収入 - 支出
- 資産 = 金融貯蓄 + 現金/預金貯蓄

給与支給日補正
- 平日：その日
- 土曜：直前の金曜
- 日曜：翌月曜
- 月間集計：補正後支給日 〜 次回補正後支給日の前日
"""

import sys
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional, Any

try:
    from PySide6.QtCore import Qt, QDate
    from PySide6.QtGui import QAction, QIcon
    from PySide6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QTabWidget, QVBoxLayout, QHBoxLayout,
        QGridLayout, QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
        QComboBox, QDateEdit, QMessageBox, QFileDialog, QSpinBox, QGroupBox,
        QHeaderView, QAbstractItemView, QDialog, QDialogButtonBox, QListWidget,
        QListWidgetItem, QInputDialog
    )
except ImportError:
    raise SystemExit(
        "PySide6 が未インストールです。\n\n"
        "pip install PySide6 openpyxl\n"
        "を実行してください。"
    )


try:
    import matplotlib
    from matplotlib import font_manager

    # matplotlibの日本語文字化け・Glyph missing警告対策。
    # Windows標準の日本語フォントを優先して登録する。
    JAPANESE_FONT_CANDIDATES = [
        (r"C:\\Windows\\Fonts\\meiryo.ttc", "Meiryo"),
        (r"C:\\Windows\\Fonts\\YuGothM.ttc", "Yu Gothic"),
        (r"C:\\Windows\\Fonts\\msgothic.ttc", "MS Gothic"),
    ]

    selected_font = None

    for font_path, font_name in JAPANESE_FONT_CANDIDATES:
        path = Path(font_path)
        if path.exists():
            try:
                font_manager.fontManager.addfont(str(path))
                selected_font = font_name
                break
            except Exception:
                pass

    if selected_font:
        matplotlib.rcParams["font.family"] = selected_font
    else:
        # フォントファイルを直接登録できない環境向けのフォールバック。
        matplotlib.rcParams["font.family"] = ["Meiryo", "Yu Gothic", "MS Gothic", "DejaVu Sans"]

    # マイナス記号が□になる問題の予防。
    matplotlib.rcParams["axes.unicode_minus"] = False

    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure

    MATPLOTLIB_AVAILABLE = True

except Exception:
    FigureCanvas = None
    Figure = None
    MATPLOTLIB_AVAILABLE = False


def get_app_dir() -> Path:
    """
    アプリの実行フォルダを返す。
    - 通常実行: .py があるフォルダ
    - EXE実行: .exe があるフォルダ

    DBはEXE内部ではなく、EXEと同じ場所へ作成する。
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def resource_path(filename: str) -> Path:
    """
    アイコンなどのリソースファイルを探す。
    優先順:
    1. EXE展開先(_MEIPASS)
    2. EXE/.py と同じフォルダ
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        bundled_path = Path(sys._MEIPASS) / filename
        if bundled_path.exists():
            return bundled_path

    return get_app_dir() / filename


APP_DIR = get_app_dir()
DB_PATH = APP_DIR / "household.db"
ICON_PATH = resource_path("kakeibo_icon.ico")
DEFAULT_SALARY_DAY = 25

DEFAULT_MASTERS = {
    "category1": ["収入", "固定費", "変動費", "特別費"],
    "category2": [
        "給与", "賞与", "住宅費", "ジム", "通信費", "AC", "生命保険",
        "クレジットカード", "iDeCo", "交通費", "娯楽費", "予備費"
    ],
    "payer": ["本人", "会社"],
}


@dataclass
class Record:
    id: Optional[int]
    category1: str
    category2: str
    income: int
    expense: int
    financial_saving: int
    cash_saving: int
    content: str
    payer: str
    tx_date: str
    memo: str


def to_int(value: Any) -> int:
    if value is None:
        return 0
    text = str(value).replace(",", "").replace("¥", "").replace("￥", "").strip()
    if text == "":
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def yen(value: int) -> str:
    return f"¥{value:,}"


def parse_date_text(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip().replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass

    return None


def adjusted_payday(year: int, month: int, salary_day: int) -> date:
    """
    給与支給日の土日補正。
    土曜は直前の金曜、日曜は翌月曜。
    """
    d = date(year, month, salary_day)

    # Monday=0 ... Saturday=5, Sunday=6
    if d.weekday() == 5:
        return d - timedelta(days=1)
    if d.weekday() == 6:
        return d + timedelta(days=1)

    return d


def monthly_period(year: int, month: int, salary_day: int = DEFAULT_SALARY_DAY) -> tuple[str, str]:
    start = adjusted_payday(year, month, salary_day)

    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1

    end = adjusted_payday(next_year, next_month, salary_day)
    return start.isoformat(), end.isoformat()


def yearly_period(year: int) -> tuple[str, str]:
    return date(year, 1, 1).isoformat(), date(year + 1, 1, 1).isoformat()


class HouseholdDB:
    def __init__(self, path: Path):
        self.path = path
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self.init_db()

    def init_db(self):
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category1 TEXT NOT NULL DEFAULT '',
                category2 TEXT NOT NULL DEFAULT '',
                income INTEGER NOT NULL DEFAULT 0,
                expense INTEGER NOT NULL DEFAULT 0,
                financial_saving INTEGER NOT NULL DEFAULT 0,
                cash_saving INTEGER NOT NULL DEFAULT 0,
                content TEXT NOT NULL DEFAULT '',
                payer TEXT NOT NULL DEFAULT '',
                tx_date TEXT NOT NULL,
                memo TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)

        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS app_config (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        """)

        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS master_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                master_type TEXT NOT NULL,
                item_name TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(master_type, item_name)
            )
        """)

        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS monthly_budgets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category2 TEXT NOT NULL UNIQUE,
                budget_amount INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)

        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS recurring_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category1 TEXT NOT NULL DEFAULT '',
                category2 TEXT NOT NULL DEFAULT '',
                income INTEGER NOT NULL DEFAULT 0,
                expense INTEGER NOT NULL DEFAULT 0,
                financial_saving INTEGER NOT NULL DEFAULT 0,
                cash_saving INTEGER NOT NULL DEFAULT 0,
                content TEXT NOT NULL DEFAULT '',
                payer TEXT NOT NULL DEFAULT '',
                day_of_month INTEGER NOT NULL DEFAULT 1,
                memo TEXT NOT NULL DEFAULT '',
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)

        self.conn.execute(
            "INSERT OR IGNORE INTO app_config (key, value) VALUES ('salary_day', ?)",
            (str(DEFAULT_SALARY_DAY),)
        )

        for master_type, items in DEFAULT_MASTERS.items():
            for index, name in enumerate(items):
                self.conn.execute("""
                    INSERT OR IGNORE INTO master_items
                    (master_type, item_name, sort_order, is_active)
                    VALUES (?, ?, ?, 1)
                """, (master_type, name, index))

        self.conn.commit()

    def get_config(self, key: str, default: str = "") -> str:
        row = self.conn.execute(
            "SELECT value FROM app_config WHERE key=?",
            (key,)
        ).fetchone()
        return str(row["value"]) if row else default

    def set_config(self, key: str, value: str):
        self.conn.execute("""
            INSERT INTO app_config (key, value)
            VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value
        """, (key, value))
        self.conn.commit()

    def get_salary_day(self) -> int:
        value = to_int(self.get_config("salary_day", str(DEFAULT_SALARY_DAY)))
        if value < 1 or value > 28:
            return DEFAULT_SALARY_DAY
        return value

    def set_salary_day(self, day: int):
        if day < 1 or day > 28:
            raise ValueError("給与支給日は1〜28の範囲で設定してください。")
        self.set_config("salary_day", str(day))

    def get_master_items(self, master_type: str) -> list[str]:
        rows = self.conn.execute("""
            SELECT item_name
            FROM master_items
            WHERE master_type=? AND is_active=1
            ORDER BY sort_order, id
        """, (master_type,)).fetchall()
        return [str(r["item_name"]) for r in rows]

    def add_master_item(self, master_type: str, item_name: str):
        item_name = item_name.strip()
        if not item_name:
            raise ValueError("追加する名称を入力してください。")

        max_order = self.conn.execute("""
            SELECT COALESCE(MAX(sort_order), -1) AS max_order
            FROM master_items
            WHERE master_type=?
        """, (master_type,)).fetchone()["max_order"]

        self.conn.execute("""
            INSERT INTO master_items (master_type, item_name, sort_order, is_active)
            VALUES (?, ?, ?, 1)
            ON CONFLICT(master_type, item_name)
            DO UPDATE SET is_active=1, updated_at=CURRENT_TIMESTAMP
        """, (master_type, item_name, int(max_order) + 1))
        self.conn.commit()

    def rename_master_item(self, master_type: str, old_name: str, new_name: str):
        old_name = old_name.strip()
        new_name = new_name.strip()

        if not old_name:
            raise ValueError("変更元が不正です。")
        if not new_name:
            raise ValueError("変更後の名称を入力してください。")

        exists = self.conn.execute("""
            SELECT id FROM master_items
            WHERE master_type=? AND item_name=? AND is_active=1
        """, (master_type, new_name)).fetchone()

        if exists and old_name != new_name:
            raise ValueError("同じ名称が既に存在します。")

        self.conn.execute("""
            UPDATE master_items
            SET item_name=?, updated_at=CURRENT_TIMESTAMP
            WHERE master_type=? AND item_name=?
        """, (new_name, master_type, old_name))
        self.conn.commit()

    def delete_master_item(self, master_type: str, item_name: str):
        """
        履歴データ側の文字列は保持し、マスタ上だけ非表示にする。
        """
        self.conn.execute("""
            UPDATE master_items
            SET is_active=0, updated_at=CURRENT_TIMESTAMP
            WHERE master_type=? AND item_name=?
        """, (master_type, item_name))
        self.conn.commit()

    def add(self, r: Record):
        self.conn.execute("""
            INSERT INTO records
            (category1, category2, income, expense, financial_saving, cash_saving,
             content, payer, tx_date, memo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            r.category1, r.category2, r.income, r.expense,
            r.financial_saving, r.cash_saving, r.content,
            r.payer, r.tx_date, r.memo
        ))
        self.conn.commit()

    def update(self, r: Record):
        if r.id is None:
            return

        self.conn.execute("""
            UPDATE records SET
                category1=?,
                category2=?,
                income=?,
                expense=?,
                financial_saving=?,
                cash_saving=?,
                content=?,
                payer=?,
                tx_date=?,
                memo=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (
            r.category1, r.category2, r.income, r.expense,
            r.financial_saving, r.cash_saving, r.content,
            r.payer, r.tx_date, r.memo, r.id
        ))
        self.conn.commit()

    def delete(self, record_id: int):
        self.conn.execute("DELETE FROM records WHERE id=?", (record_id,))
        self.conn.commit()

    def list_all(self):
        return self.conn.execute("""
            SELECT *
            FROM records
            ORDER BY tx_date, id
        """).fetchall()

    def list_between(self, start: str, end: str):
        return self.conn.execute("""
            SELECT *
            FROM records
            WHERE tx_date >= ? AND tx_date < ?
            ORDER BY tx_date, id
        """, (start, end)).fetchall()

    def summary_between(self, start: Optional[str] = None, end: Optional[str] = None):
        where = ""
        params: tuple[Any, ...] = ()

        if start and end:
            where = "WHERE tx_date >= ? AND tx_date < ?"
            params = (start, end)

        row = self.conn.execute(f"""
            SELECT
                COALESCE(SUM(income), 0) AS income,
                COALESCE(SUM(expense), 0) AS expense,
                COALESCE(SUM(financial_saving), 0) AS financial_saving,
                COALESCE(SUM(cash_saving), 0) AS cash_saving
            FROM records
            {where}
        """, params).fetchone()

        income = int(row["income"])
        expense_only = int(row["expense"])
        financial = int(row["financial_saving"])
        cash = int(row["cash_saving"])
        expense_total = expense_only + financial + cash

        return {
            "income": income,
            "expense_only": expense_only,
            "expense_total": expense_total,
            "financial_saving": financial,
            "cash_saving": cash,
            "profit": income - expense_total,
            "assets": financial + cash,
        }

    def get_monthly_budgets(self):
        return self.conn.execute("""
            SELECT *
            FROM monthly_budgets
            WHERE is_active=1
            ORDER BY category2
        """).fetchall()

    def upsert_monthly_budget(self, category2: str, amount: int):
        category2 = category2.strip()
        if not category2:
            raise ValueError("中区分を入力してください。")
        if amount < 0:
            raise ValueError("予算金額は0以上で入力してください。")

        self.conn.execute("""
            INSERT INTO monthly_budgets (category2, budget_amount, is_active)
            VALUES (?, ?, 1)
            ON CONFLICT(category2)
            DO UPDATE SET budget_amount=excluded.budget_amount,
                          is_active=1,
                          updated_at=CURRENT_TIMESTAMP
        """, (category2, amount))
        self.conn.commit()

    def delete_monthly_budget(self, category2: str):
        self.conn.execute("""
            UPDATE monthly_budgets
            SET is_active=0, updated_at=CURRENT_TIMESTAMP
            WHERE category2=?
        """, (category2,))
        self.conn.commit()

    def category_expense_between(self, start: str, end: str):
        return self.conn.execute("""
            SELECT
                category2,
                COALESCE(SUM(expense + financial_saving + cash_saving), 0) AS total_expense
            FROM records
            WHERE tx_date >= ? AND tx_date < ?
            GROUP BY category2
            HAVING total_expense > 0
            ORDER BY total_expense DESC
        """, (start, end)).fetchall()

    def monthly_profit_series(self, year: int, salary_day: int):
        data = []
        for month in range(1, 13):
            start, end = monthly_period(year, month, salary_day)
            s = self.summary_between(start, end)
            data.append((month, s["profit"], s["income"], s["expense_total"], s["assets"]))
        return data

    def get_recurring_items(self):
        return self.conn.execute("""
            SELECT *
            FROM recurring_items
            WHERE is_active=1
            ORDER BY day_of_month, id
        """).fetchall()

    def add_recurring_item(self, r: Record, day_of_month: int):
        if day_of_month < 1 or day_of_month > 28:
            raise ValueError("定期支出の日付は1〜28で入力してください。")

        self.conn.execute("""
            INSERT INTO recurring_items
            (category1, category2, income, expense, financial_saving, cash_saving,
             content, payer, day_of_month, memo, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
        """, (
            r.category1, r.category2, r.income, r.expense, r.financial_saving,
            r.cash_saving, r.content, r.payer, day_of_month, r.memo
        ))
        self.conn.commit()

    def delete_recurring_item(self, item_id: int):
        self.conn.execute("""
            UPDATE recurring_items
            SET is_active=0, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (item_id,))
        self.conn.commit()

    def generate_recurring_for_month(self, year: int, month: int) -> int:
        items = self.get_recurring_items()
        count = 0

        for item in items:
            tx_date = date(year, month, int(item["day_of_month"])).isoformat()

            exists = self.conn.execute("""
                SELECT id
                FROM records
                WHERE tx_date=?
                  AND category1=?
                  AND category2=?
                  AND content=?
                  AND expense=?
                  AND financial_saving=?
                  AND cash_saving=?
                  AND memo LIKE '%[定期作成]%'
                LIMIT 1
            """, (
                tx_date,
                item["category1"],
                item["category2"],
                item["content"],
                int(item["expense"]),
                int(item["financial_saving"]),
                int(item["cash_saving"]),
            )).fetchone()

            if exists:
                continue

            memo = str(item["memo"] or "")
            if memo:
                memo = memo + " / [定期作成]"
            else:
                memo = "[定期作成]"

            self.conn.execute("""
                INSERT INTO records
                (category1, category2, income, expense, financial_saving, cash_saving,
                 content, payer, tx_date, memo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                item["category1"],
                item["category2"],
                int(item["income"]),
                int(item["expense"]),
                int(item["financial_saving"]),
                int(item["cash_saving"]),
                item["content"],
                item["payer"],
                tx_date,
                memo,
            ))
            count += 1

        self.conn.commit()
        return count

    def import_excel(self, excel_path: str) -> int:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("Excel取込には openpyxl が必要です。pip install openpyxl を実行してください。")

        wb = load_workbook(excel_path, data_only=True, keep_vba=True)

        if "家計簿" not in wb.sheetnames:
            raise RuntimeError("Excel内に『家計簿』シートが見つかりません。")

        ws = wb["家計簿"]
        count = 0

        for row in range(6, ws.max_row + 1):
            tx_date = parse_date_text(ws.cell(row, 10).value)
            if not tx_date:
                continue

            category1 = str(ws.cell(row, 2).value or "").strip()
            category2 = str(ws.cell(row, 3).value or "").strip()

            if not category1 and not category2:
                continue

            self.add(Record(
                id=None,
                category1=category1,
                category2=category2,
                income=to_int(ws.cell(row, 4).value),
                expense=to_int(ws.cell(row, 5).value),
                financial_saving=to_int(ws.cell(row, 6).value),
                cash_saving=to_int(ws.cell(row, 7).value),
                content=str(ws.cell(row, 8).value or "").strip(),
                payer=str(ws.cell(row, 9).value or "").strip(),
                tx_date=tx_date,
                memo=str(ws.cell(row, 15).value or "").strip(),
            ))
            count += 1

            # Excelから拾った値もマスタに登録しておく
            if category1:
                try:
                    self.add_master_item("category1", category1)
                except Exception:
                    pass

            if category2:
                try:
                    self.add_master_item("category2", category2)
                except Exception:
                    pass

            payer = str(ws.cell(row, 9).value or "").strip()
            if payer:
                try:
                    self.add_master_item("payer", payer)
                except Exception:
                    pass

        return count


class SummaryCard(QGroupBox):
    def __init__(self, title: str):
        super().__init__(title)
        self.label = QLabel("¥0")
        self.label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        layout = QVBoxLayout(self)
        layout.addWidget(self.label)

        self.set_yen(0)

    def set_yen(self, value: int):
        self.label.setText(yen(value))
        color = "#d32f2f" if value < 0 else "#111827"
        self.label.setStyleSheet(
            f"font-size:24px;font-weight:bold;padding:12px;color:{color};"
        )


class SalaryDayConfigDialog(QDialog):
    def __init__(self, current_day: int, parent=None):
        super().__init__(parent)
        self.setWindowTitle("コンフィグ - 給与支給日設定")
        self.resize(520, 200)

        layout = QVBoxLayout(self)

        description = QLabel(
            "給与支給日を設定します。\n"
            "月間集計は『補正後の支給日〜次回の補正後支給日の前日』で集計します。\n"
            "土曜の場合は直前の金曜、日曜の場合は翌月曜に補正します。"
        )
        description.setWordWrap(True)
        layout.addWidget(description)

        row = QHBoxLayout()
        row.addWidget(QLabel("給与支給日"))

        self.day_spin = QSpinBox()
        self.day_spin.setRange(1, 28)
        self.day_spin.setValue(current_day)

        row.addWidget(self.day_spin)
        row.addWidget(QLabel("日"))
        row.addStretch()
        layout.addLayout(row)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def salary_day(self) -> int:
        return self.day_spin.value()


class MasterConfigDialog(QDialog):
    MASTER_LABELS = {
        "category1": "大区分",
        "category2": "中区分",
        "payer": "支払者",
    }

    def __init__(self, db: HouseholdDB, master_type: str, parent=None):
        super().__init__(parent)
        self.db = db
        self.master_type = master_type
        self.label_name = self.MASTER_LABELS.get(master_type, master_type)

        self.setWindowTitle(f"コンフィグ - {self.label_name}設定")
        self.resize(520, 520)

        layout = QVBoxLayout(self)

        description = QLabel(
            f"{self.label_name}のプルダウン候補を管理します。\n"
            "ここで追加・編集・削除した内容は入力画面のリストへ反映されます。\n"
            "削除しても過去の家計簿データの文字列は消えません。"
        )
        description.setWordWrap(True)
        layout.addWidget(description)

        self.list_widget = QListWidget()
        layout.addWidget(self.list_widget)

        btn_row = QHBoxLayout()

        add_btn = QPushButton("追加")
        edit_btn = QPushButton("編集")
        delete_btn = QPushButton("削除")
        close_btn = QPushButton("閉じる")

        add_btn.clicked.connect(self.add_item)
        edit_btn.clicked.connect(self.edit_item)
        delete_btn.clicked.connect(self.delete_item)
        close_btn.clicked.connect(self.accept)

        btn_row.addWidget(add_btn)
        btn_row.addWidget(edit_btn)
        btn_row.addWidget(delete_btn)
        btn_row.addStretch()
        btn_row.addWidget(close_btn)

        layout.addLayout(btn_row)

        self.reload_items()

    def reload_items(self):
        self.list_widget.clear()
        for name in self.db.get_master_items(self.master_type):
            self.list_widget.addItem(QListWidgetItem(name))

    def selected_name(self) -> Optional[str]:
        item = self.list_widget.currentItem()
        if item is None:
            return None
        return item.text()

    def add_item(self):
        text, ok = QInputDialog.getText(
            self,
            f"{self.label_name}追加",
            f"追加する{self.label_name}名を入力してください。"
        )

        if not ok:
            return

        try:
            self.db.add_master_item(self.master_type, text)
        except Exception as e:
            QMessageBox.warning(self, "追加エラー", str(e))
            return

        self.reload_items()

    def edit_item(self):
        old_name = self.selected_name()
        if old_name is None:
            QMessageBox.warning(self, "編集エラー", "編集する項目を選択してください。")
            return

        text, ok = QInputDialog.getText(
            self,
            f"{self.label_name}編集",
            f"{self.label_name}名を変更してください。",
            text=old_name
        )

        if not ok:
            return

        try:
            self.db.rename_master_item(self.master_type, old_name, text)
        except Exception as e:
            QMessageBox.warning(self, "編集エラー", str(e))
            return

        self.reload_items()

    def delete_item(self):
        name = self.selected_name()
        if name is None:
            QMessageBox.warning(self, "削除エラー", "削除する項目を選択してください。")
            return

        reply = QMessageBox.question(
            self,
            "削除確認",
            f"『{name}』を{self.label_name}リストから削除しますか？\n"
            "過去の家計簿データは削除されません。"
        )

        if reply != QMessageBox.Yes:
            return

        self.db.delete_master_item(self.master_type, name)
        self.reload_items()


class BudgetConfigDialog(QDialog):
    def __init__(self, db: HouseholdDB, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("コンフィグ - 月予算設定")
        self.resize(680, 520)

        layout = QVBoxLayout(self)

        note = QLabel(
            "中区分ごとの月予算を設定します。\n"
            "月間収益画面では、実績が予算を超過した項目を赤表示します。"
        )
        note.setWordWrap(True)
        layout.addWidget(note)

        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["中区分", "月予算"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("追加")
        save_btn = QPushButton("保存")
        delete_btn = QPushButton("選択行を削除")
        close_btn = QPushButton("閉じる")

        add_btn.clicked.connect(self.add_row)
        save_btn.clicked.connect(self.save)
        delete_btn.clicked.connect(self.delete_selected)
        close_btn.clicked.connect(self.accept)

        btn_row.addWidget(add_btn)
        btn_row.addWidget(save_btn)
        btn_row.addWidget(delete_btn)
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        self.reload()

    def reload(self):
        rows = self.db.get_monthly_budgets()
        self.table.setRowCount(len(rows))

        for i, row in enumerate(rows):
            self.table.setItem(i, 0, QTableWidgetItem(str(row["category2"])))
            amount_item = QTableWidgetItem(str(int(row["budget_amount"])))
            amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(i, 1, amount_item)

    def add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, QTableWidgetItem(""))
        self.table.setItem(row, 1, QTableWidgetItem("0"))

    def save(self):
        try:
            for row in range(self.table.rowCount()):
                category_item = self.table.item(row, 0)
                amount_item = self.table.item(row, 1)

                category2 = category_item.text().strip() if category_item else ""
                amount = to_int(amount_item.text() if amount_item else 0)

                if not category2:
                    continue

                self.db.upsert_monthly_budget(category2, amount)
                self.db.add_master_item("category2", category2)

        except Exception as e:
            QMessageBox.warning(self, "保存エラー", str(e))
            return

        self.reload()
        QMessageBox.information(self, "保存完了", "月予算を保存しました。")

    def delete_selected(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "削除エラー", "削除する行を選択してください。")
            return

        item = self.table.item(row, 0)
        category2 = item.text().strip() if item else ""

        if not category2:
            self.table.removeRow(row)
            return

        reply = QMessageBox.question(
            self,
            "削除確認",
            f"『{category2}』の月予算設定を削除しますか？"
        )

        if reply != QMessageBox.Yes:
            return

        self.db.delete_monthly_budget(category2)
        self.reload()


class RecurringConfigDialog(QDialog):
    def __init__(self, db: HouseholdDB, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("コンフィグ - 定期支出登録")
        self.resize(1100, 600)

        layout = QVBoxLayout(self)

        note = QLabel(
            "毎月発生する固定費・積立などを登録します。\n"
            "『指定月へ作成』を押すと、指定した年月の家計簿データとして自動作成します。"
        )
        note.setWordWrap(True)
        layout.addWidget(note)

        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "ID", "日", "大区分", "中区分", "収入", "支出", "金融貯蓄",
            "現金/預金貯蓄", "内容", "支払者"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)

        add_box = QGroupBox("新規定期項目")
        form = QGridLayout(add_box)

        self.day_spin = QSpinBox()
        self.day_spin.setRange(1, 28)
        self.day_spin.setValue(25)

        self.category1 = QComboBox()
        self.category1.setEditable(True)
        self.category1.addItems(self.db.get_master_items("category1"))

        self.category2 = QComboBox()
        self.category2.setEditable(True)
        self.category2.addItems(self.db.get_master_items("category2"))

        self.income = QLineEdit()
        self.expense = QLineEdit()
        self.financial = QLineEdit()
        self.cash = QLineEdit()
        self.content = QLineEdit()

        self.payer = QComboBox()
        self.payer.setEditable(True)
        self.payer.addItems(self.db.get_master_items("payer"))

        fields = [
            ("日", self.day_spin),
            ("大区分", self.category1),
            ("中区分", self.category2),
            ("収入", self.income),
            ("支出", self.expense),
            ("金融貯蓄", self.financial),
            ("現金/預金貯蓄", self.cash),
            ("内容", self.content),
            ("支払者", self.payer),
        ]

        for i, (label, widget) in enumerate(fields):
            form.addWidget(QLabel(label), i // 3, (i % 3) * 2)
            form.addWidget(widget, i // 3, (i % 3) * 2 + 1)

        layout.addWidget(add_box)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("定期項目を追加")
        delete_btn = QPushButton("選択行を削除")
        generate_btn = QPushButton("指定月へ作成")
        close_btn = QPushButton("閉じる")

        add_btn.clicked.connect(self.add_recurring)
        delete_btn.clicked.connect(self.delete_selected)
        generate_btn.clicked.connect(self.generate_for_month)
        close_btn.clicked.connect(self.accept)

        self.year_combo = QComboBox()
        self.year_combo.addItems([str(y) for y in range(2000, 2101)])
        self.year_combo.setCurrentText(str(date.today().year))

        self.month_combo = QComboBox()
        self.month_combo.addItems([str(m) for m in range(1, 13)])
        self.month_combo.setCurrentText(str(date.today().month))

        btn_row.addWidget(add_btn)
        btn_row.addWidget(delete_btn)
        btn_row.addStretch()
        btn_row.addWidget(QLabel("作成年"))
        btn_row.addWidget(self.year_combo)
        btn_row.addWidget(QLabel("月"))
        btn_row.addWidget(self.month_combo)
        btn_row.addWidget(generate_btn)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        self.reload()

    def reload(self):
        rows = self.db.get_recurring_items()
        self.table.setRowCount(len(rows))

        for i, row in enumerate(rows):
            values = [
                row["id"],
                row["day_of_month"],
                row["category1"],
                row["category2"],
                yen(row["income"]) if row["income"] else "",
                yen(row["expense"]) if row["expense"] else "",
                yen(row["financial_saving"]) if row["financial_saving"] else "",
                yen(row["cash_saving"]) if row["cash_saving"] else "",
                row["content"],
                row["payer"],
            ]

            for col, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if col in (4, 5, 6, 7):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(i, col, item)

    def add_recurring(self):
        r = Record(
            id=None,
            category1=self.category1.currentText().strip(),
            category2=self.category2.currentText().strip(),
            income=to_int(self.income.text()),
            expense=to_int(self.expense.text()),
            financial_saving=to_int(self.financial.text()),
            cash_saving=to_int(self.cash.text()),
            content=self.content.text().strip(),
            payer=self.payer.currentText().strip(),
            tx_date=date.today().isoformat(),
            memo="",
        )

        if not r.category1 or not r.category2:
            QMessageBox.warning(self, "入力エラー", "大区分と中区分を入力してください。")
            return

        if r.income == 0 and r.expense == 0 and r.financial_saving == 0 and r.cash_saving == 0:
            QMessageBox.warning(self, "入力エラー", "金額を入力してください。")
            return

        try:
            self.db.add_recurring_item(r, self.day_spin.value())
            self.db.add_master_item("category1", r.category1)
            self.db.add_master_item("category2", r.category2)
            if r.payer:
                self.db.add_master_item("payer", r.payer)
        except Exception as e:
            QMessageBox.warning(self, "登録エラー", str(e))
            return

        self.income.clear()
        self.expense.clear()
        self.financial.clear()
        self.cash.clear()
        self.content.clear()
        self.reload()

    def delete_selected(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "削除エラー", "削除する行を選択してください。")
            return

        item = self.table.item(row, 0)
        if item is None:
            return

        item_id = int(item.text())

        reply = QMessageBox.question(
            self,
            "削除確認",
            f"定期項目 ID {item_id} を削除しますか？"
        )

        if reply != QMessageBox.Yes:
            return

        self.db.delete_recurring_item(item_id)
        self.reload()

    def generate_for_month(self):
        y = int(self.year_combo.currentText())
        m = int(self.month_combo.currentText())

        reply = QMessageBox.question(
            self,
            "定期支出作成確認",
            f"{y}年{m}月分の定期項目を家計簿へ作成しますか？\n"
            "同じ定期作成データが既にある場合は重複作成しません。"
        )

        if reply != QMessageBox.Yes:
            return

        try:
            count = self.db.generate_recurring_for_month(y, m)
        except Exception as e:
            QMessageBox.critical(self, "作成エラー", str(e))
            return

        QMessageBox.information(self, "作成完了", f"{count}件作成しました。")


def apply_app_icon(target):
    """
    ウィンドウ・アプリケーションにアイコンを設定する。
    kakeibo_icon.ico が存在する場合のみ反映する。
    """
    if ICON_PATH.exists():
        target.setWindowIcon(QIcon(str(ICON_PATH)))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.db = HouseholdDB(DB_PATH)
        self.salary_day = self.db.get_salary_day()
        self.current_rows = []
        self.editing_id: Optional[int] = None

        self.setWindowTitle("家計簿 GUI版 v0.7")
        self.resize(1400, 850)
        apply_app_icon(self)

        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        self.build_menu()
        self.build_dashboard_tab()
        self.build_input_tab()
        self.build_list_tab()
        self.build_monthly_tab()
        self.build_yearly_tab()
        self.build_graph_tab()

        self.refresh_master_combos()
        self.refresh_all()

    def build_menu(self):
        file_menu = self.menuBar().addMenu("ファイル")

        import_action = QAction("Excelから取込", self)
        import_action.triggered.connect(self.import_excel)
        file_menu.addAction(import_action)

        exit_action = QAction("終了", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        config_menu = self.menuBar().addMenu("コンフィグ")

        category1_action = QAction("大区分設定", self)
        category1_action.triggered.connect(lambda: self.open_master_config("category1"))
        config_menu.addAction(category1_action)

        category2_action = QAction("中区分設定", self)
        category2_action.triggered.connect(lambda: self.open_master_config("category2"))
        config_menu.addAction(category2_action)

        payer_action = QAction("支払者設定", self)
        payer_action.triggered.connect(lambda: self.open_master_config("payer"))
        config_menu.addAction(payer_action)

        config_menu.addSeparator()

        budget_action = QAction("月予算設定", self)
        budget_action.triggered.connect(self.open_budget_config)
        config_menu.addAction(budget_action)

        recurring_action = QAction("定期支出登録", self)
        recurring_action.triggered.connect(self.open_recurring_config)
        config_menu.addAction(recurring_action)

        config_menu.addSeparator()

        salary_action = QAction("給与支給日設定", self)
        salary_action.triggered.connect(self.open_salary_day_config)
        config_menu.addAction(salary_action)

    def build_dashboard_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        self.total_income = SummaryCard("全期間 収入")
        self.total_expense = SummaryCard("全期間 支出（支出＋貯蓄）")
        self.total_profit = SummaryCard("全期間 収益")
        self.total_assets = SummaryCard("全期間 資産")

        grid = QGridLayout()
        grid.addWidget(self.total_income, 0, 0)
        grid.addWidget(self.total_expense, 0, 1)
        grid.addWidget(self.total_profit, 1, 0)
        grid.addWidget(self.total_assets, 1, 1)

        layout.addLayout(grid)
        layout.addStretch()

        self.tabs.addTab(tab, "総収益")

    def build_input_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        form = QGridLayout()

        self.category1 = QComboBox()
        self.category1.setEditable(True)

        self.category2 = QComboBox()
        self.category2.setEditable(True)

        self.income = QLineEdit()
        self.expense = QLineEdit()
        self.financial = QLineEdit()
        self.cash = QLineEdit()
        self.content = QLineEdit()

        self.payer = QComboBox()
        self.payer.setEditable(True)

        self.tx_date = QDateEdit()
        self.tx_date.setCalendarPopup(True)
        self.tx_date.setDate(QDate.currentDate())

        self.memo = QLineEdit()

        fields = [
            ("大区分", self.category1),
            ("中区分", self.category2),
            ("収入", self.income),
            ("支出", self.expense),
            ("金融貯蓄", self.financial),
            ("現金/預金貯蓄", self.cash),
            ("内容", self.content),
            ("支払者", self.payer),
            ("入出日", self.tx_date),
            ("メモ", self.memo),
        ]

        for row, (label, widget) in enumerate(fields):
            form.addWidget(QLabel(label), row, 0)
            form.addWidget(widget, row, 1)

        btns = QHBoxLayout()

        self.add_btn = QPushButton("登録")
        self.update_btn = QPushButton("更新")
        self.clear_btn = QPushButton("クリア")

        self.add_btn.clicked.connect(self.add_record)
        self.update_btn.clicked.connect(self.update_record)
        self.clear_btn.clicked.connect(self.clear_form)

        btns.addWidget(self.add_btn)
        btns.addWidget(self.update_btn)
        btns.addWidget(self.clear_btn)

        layout.addLayout(form)
        layout.addLayout(btns)
        layout.addStretch()

        self.tabs.addTab(tab, "入力")

    def build_list_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        controls = QHBoxLayout()

        self.filter_year = QComboBox()
        self.filter_year.addItems([str(y) for y in range(2000, 2101)])
        self.filter_year.setCurrentText(str(date.today().year))

        self.filter_month = QComboBox()
        self.filter_month.addItems([str(m) for m in range(1, 13)])
        self.filter_month.setCurrentText(str(date.today().month))

        btn_month = QPushButton("更新（月間表示）")
        btn_all = QPushButton("全期間表示")
        btn_year = QPushButton("年間表示")

        btn_month.clicked.connect(self.show_month_records)
        btn_all.clicked.connect(self.show_all_records)
        btn_year.clicked.connect(self.show_year_records)

        controls.addWidget(QLabel("年"))
        controls.addWidget(self.filter_year)
        controls.addWidget(QLabel("月"))
        controls.addWidget(self.filter_month)
        controls.addWidget(btn_month)
        controls.addWidget(btn_all)
        controls.addWidget(btn_year)
        controls.addStretch()

        self.table = QTableWidget()
        self.table.setColumnCount(14)
        self.table.setHorizontalHeaderLabels([
            "ID", "大区分", "中区分", "収入", "支出", "金融貯蓄",
            "現金/預金貯蓄", "内容", "支払者", "入出日", "年", "月", "曜日", "メモ"
        ])

        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.doubleClicked.connect(self.load_selected_to_form)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        delete_btn = QPushButton("選択行を削除")
        delete_btn.clicked.connect(self.delete_selected)

        layout.addLayout(controls)
        layout.addWidget(self.table)
        layout.addWidget(delete_btn)

        self.tabs.addTab(tab, "家計簿一覧")

    def build_monthly_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        row = QHBoxLayout()

        self.month_summary_year = QSpinBox()
        self.month_summary_year.setRange(2000, 2100)
        self.month_summary_year.setValue(date.today().year)

        self.month_summary_month = QSpinBox()
        self.month_summary_month.setRange(1, 12)
        self.month_summary_month.setValue(date.today().month)

        btn = QPushButton("月間集計")
        btn.clicked.connect(self.refresh_monthly)

        row.addWidget(QLabel("年"))
        row.addWidget(self.month_summary_year)
        row.addWidget(QLabel("月"))
        row.addWidget(self.month_summary_month)
        row.addWidget(btn)
        row.addStretch()

        self.month_period_label = QLabel("")

        self.month_cards = {
            "income": SummaryCard("収入"),
            "expense_total": SummaryCard("支出（支出＋金融貯蓄＋現金/預金貯蓄）"),
            "financial_saving": SummaryCard("金融貯蓄"),
            "cash_saving": SummaryCard("現金/預金貯蓄"),
            "profit": SummaryCard("収益"),
            "assets": SummaryCard("資産"),
        }

        grid = QGridLayout()
        for i, card in enumerate(self.month_cards.values()):
            grid.addWidget(card, i // 2, i % 2)

        self.budget_table = QTableWidget()
        self.budget_table.setColumnCount(5)
        self.budget_table.setHorizontalHeaderLabels(["中区分", "実績", "予算", "差額", "判定"])
        self.budget_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.budget_table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        layout.addLayout(row)
        layout.addWidget(self.month_period_label)
        layout.addLayout(grid)
        layout.addWidget(QLabel("月予算チェック"))
        layout.addWidget(self.budget_table)
        layout.addStretch()

        self.tabs.addTab(tab, "月間収益")

    def build_yearly_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        row = QHBoxLayout()

        self.year_summary_year = QSpinBox()
        self.year_summary_year.setRange(2000, 2100)
        self.year_summary_year.setValue(date.today().year)

        btn = QPushButton("年間集計")
        btn.clicked.connect(self.refresh_yearly)

        row.addWidget(QLabel("年"))
        row.addWidget(self.year_summary_year)
        row.addWidget(btn)
        row.addStretch()

        self.year_period_label = QLabel("")

        self.year_cards = {
            "income": SummaryCard("収入"),
            "expense_total": SummaryCard("支出（支出＋金融貯蓄＋現金/預金貯蓄）"),
            "financial_saving": SummaryCard("金融貯蓄"),
            "cash_saving": SummaryCard("現金/預金貯蓄"),
            "profit": SummaryCard("収益"),
            "assets": SummaryCard("資産"),
        }

        grid = QGridLayout()
        for i, card in enumerate(self.year_cards.values()):
            grid.addWidget(card, i // 2, i % 2)

        layout.addLayout(row)
        layout.addWidget(self.year_period_label)
        layout.addLayout(grid)
        layout.addStretch()

        self.tabs.addTab(tab, "年間収益")

    def build_graph_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        controls = QHBoxLayout()

        self.graph_year = QComboBox()
        self.graph_year.addItems([str(y) for y in range(2000, 2101)])
        self.graph_year.setCurrentText(str(date.today().year))

        self.graph_month = QComboBox()
        self.graph_month.addItems([str(m) for m in range(1, 13)])
        self.graph_month.setCurrentText(str(date.today().month))

        refresh_btn = QPushButton("グラフ更新")
        refresh_btn.clicked.connect(self.refresh_graphs)

        controls.addWidget(QLabel("年"))
        controls.addWidget(self.graph_year)
        controls.addWidget(QLabel("月"))
        controls.addWidget(self.graph_month)
        controls.addWidget(refresh_btn)
        controls.addStretch()

        layout.addLayout(controls)

        if MATPLOTLIB_AVAILABLE:
            self.graph_figure = Figure(figsize=(10, 7))
            self.graph_canvas = FigureCanvas(self.graph_figure)
            layout.addWidget(self.graph_canvas)
        else:
            self.graph_figure = None
            self.graph_canvas = None
            label = QLabel(
                "matplotlib が未インストールのため、グラフを表示できません。\n"
                "以下を実行してください。\n\n"
                "pip install matplotlib"
            )
            layout.addWidget(label)

        self.tabs.addTab(tab, "分析グラフ")

    def refresh_master_combos(self):
        current_category1 = self.category1.currentText() if hasattr(self, "category1") else ""
        current_category2 = self.category2.currentText() if hasattr(self, "category2") else ""
        current_payer = self.payer.currentText() if hasattr(self, "payer") else ""

        self.category1.clear()
        self.category1.addItems(self.db.get_master_items("category1"))
        if current_category1:
            self.category1.setCurrentText(current_category1)

        self.category2.clear()
        self.category2.addItems(self.db.get_master_items("category2"))
        if current_category2:
            self.category2.setCurrentText(current_category2)

        self.payer.clear()
        self.payer.addItems(self.db.get_master_items("payer"))
        if current_payer:
            self.payer.setCurrentText(current_payer)
        elif self.payer.count() > 0:
            self.payer.setCurrentIndex(0)

    def form_record(self) -> Record:
        qd = self.tx_date.date()

        return Record(
            id=self.editing_id,
            category1=self.category1.currentText().strip(),
            category2=self.category2.currentText().strip(),
            income=to_int(self.income.text()),
            expense=to_int(self.expense.text()),
            financial_saving=to_int(self.financial.text()),
            cash_saving=to_int(self.cash.text()),
            content=self.content.text().strip(),
            payer=self.payer.currentText().strip(),
            tx_date=date(qd.year(), qd.month(), qd.day()).isoformat(),
            memo=self.memo.text().strip(),
        )

    def validate_record(self, r: Record) -> bool:
        if not r.category1:
            QMessageBox.warning(self, "入力エラー", "大区分を入力してください。")
            return False

        if not r.category2:
            QMessageBox.warning(self, "入力エラー", "中区分を入力してください。")
            return False

        if r.income == 0 and r.expense == 0 and r.financial_saving == 0 and r.cash_saving == 0:
            QMessageBox.warning(self, "入力エラー", "金額を入力してください。")
            return False

        return True

    def add_record(self):
        r = self.form_record()

        if not self.validate_record(r):
            return

        # 手入力された候補もマスタへ自動追加
        for master_type, value in [
            ("category1", r.category1),
            ("category2", r.category2),
            ("payer", r.payer),
        ]:
            if value:
                try:
                    self.db.add_master_item(master_type, value)
                except Exception:
                    pass

        self.db.add(r)
        self.clear_form()
        self.refresh_master_combos()
        self.refresh_all()

        QMessageBox.information(self, "登録完了", "登録しました。")

    def update_record(self):
        if self.editing_id is None:
            QMessageBox.warning(self, "更新エラー", "一覧から編集対象をダブルクリックしてください。")
            return

        r = self.form_record()

        if not self.validate_record(r):
            return

        for master_type, value in [
            ("category1", r.category1),
            ("category2", r.category2),
            ("payer", r.payer),
        ]:
            if value:
                try:
                    self.db.add_master_item(master_type, value)
                except Exception:
                    pass

        self.db.update(r)
        self.clear_form()
        self.refresh_master_combos()
        self.refresh_all()

        QMessageBox.information(self, "更新完了", "更新しました。")

    def clear_form(self):
        self.editing_id = None

        if self.category1.count() > 0:
            self.category1.setCurrentIndex(0)
        else:
            self.category1.setCurrentText("")

        if self.category2.count() > 0:
            self.category2.setCurrentIndex(0)
        else:
            self.category2.setCurrentText("")

        self.income.clear()
        self.expense.clear()
        self.financial.clear()
        self.cash.clear()
        self.content.clear()
        self.memo.clear()

        if self.payer.count() > 0:
            self.payer.setCurrentIndex(0)
        else:
            self.payer.setCurrentText("")

        self.tx_date.setDate(QDate.currentDate())

    def selected_record_id(self) -> Optional[int]:
        row = self.table.currentRow()

        if row < 0:
            return None

        item = self.table.item(row, 0)

        if item is None:
            return None

        return int(item.text())

    def load_selected_to_form(self):
        rid = self.selected_record_id()

        if rid is None:
            return

        for r in self.current_rows:
            if int(r["id"]) == rid:
                self.editing_id = rid

                self.category1.setCurrentText(r["category1"])
                self.category2.setCurrentText(r["category2"])
                self.income.setText(str(r["income"] or ""))
                self.expense.setText(str(r["expense"] or ""))
                self.financial.setText(str(r["financial_saving"] or ""))
                self.cash.setText(str(r["cash_saving"] or ""))
                self.content.setText(r["content"])
                self.payer.setCurrentText(r["payer"])
                self.memo.setText(r["memo"])

                d = datetime.strptime(r["tx_date"], "%Y-%m-%d").date()
                self.tx_date.setDate(QDate(d.year, d.month, d.day))

                self.tabs.setCurrentIndex(1)
                return

    def delete_selected(self):
        rid = self.selected_record_id()

        if rid is None:
            QMessageBox.warning(self, "削除エラー", "削除する行を選択してください。")
            return

        reply = QMessageBox.question(
            self,
            "削除確認",
            f"ID {rid} を削除しますか？"
        )

        if reply == QMessageBox.Yes:
            self.db.delete(rid)
            self.refresh_all()

    def set_table_rows(self, rows):
        self.current_rows = rows
        self.table.setRowCount(len(rows))

        weekdays = ["月", "火", "水", "木", "金", "土", "日"]

        for row_index, r in enumerate(rows):
            d = datetime.strptime(r["tx_date"], "%Y-%m-%d").date()

            values = [
                r["id"],
                r["category1"],
                r["category2"],
                yen(r["income"]) if r["income"] else "",
                yen(r["expense"]) if r["expense"] else "",
                yen(r["financial_saving"]) if r["financial_saving"] else "",
                yen(r["cash_saving"]) if r["cash_saving"] else "",
                r["content"],
                r["payer"],
                r["tx_date"],
                d.year,
                d.month,
                weekdays[d.weekday()],
                r["memo"],
            ]

            for col, value in enumerate(values):
                item = QTableWidgetItem(str(value))

                if col in (3, 4, 5, 6):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

                self.table.setItem(row_index, col, item)

    def show_all_records(self):
        self.set_table_rows(self.db.list_all())

    def show_month_records(self):
        start, end = monthly_period(
            int(self.filter_year.currentText()),
            int(self.filter_month.currentText()),
            self.salary_day
        )
        self.set_table_rows(self.db.list_between(start, end))

    def show_year_records(self):
        start, end = yearly_period(int(self.filter_year.currentText()))
        self.set_table_rows(self.db.list_between(start, end))

    def refresh_dashboard(self):
        s = self.db.summary_between()

        self.total_income.set_yen(s["income"])
        self.total_expense.set_yen(s["expense_total"])
        self.total_profit.set_yen(s["profit"])
        self.total_assets.set_yen(s["assets"])

    def refresh_monthly(self):
        y = self.month_summary_year.value()
        m = self.month_summary_month.value()

        start, end = monthly_period(y, m, self.salary_day)
        s = self.db.summary_between(start, end)

        self.month_period_label.setText(
            f"給与支給日設定：{self.salary_day}日 / 集計期間：{start} 〜 {end} の前日まで"
        )

        for key, card in self.month_cards.items():
            card.set_yen(s[key])

        self.refresh_budget_table(start, end)

    def refresh_budget_table(self, start: str, end: str):
        budgets = self.db.get_monthly_budgets()
        actual_rows = self.db.category_expense_between(start, end)
        actual_map = {
            str(r["category2"]): int(r["total_expense"])
            for r in actual_rows
        }

        self.budget_table.setRowCount(len(budgets))

        for row_index, budget in enumerate(budgets):
            category2 = str(budget["category2"])
            budget_amount = int(budget["budget_amount"])
            actual = actual_map.get(category2, 0)
            diff = budget_amount - actual
            status = "OK" if actual <= budget_amount else "超過"

            values = [
                category2,
                yen(actual),
                yen(budget_amount),
                yen(diff),
                status,
            ]

            for col, value in enumerate(values):
                item = QTableWidgetItem(str(value))

                if col in (1, 2, 3):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

                if status == "超過":
                    item.setForeground(Qt.GlobalColor.red)

                self.budget_table.setItem(row_index, col, item)

    def refresh_yearly(self):
        y = self.year_summary_year.value()

        start, end = yearly_period(y)
        s = self.db.summary_between(start, end)

        self.year_period_label.setText(f"集計期間：{y}/1/1 〜 {y}/12/31")

        for key, card in self.year_cards.items():
            card.set_yen(s[key])

    def refresh_all(self):
        self.refresh_dashboard()
        self.refresh_monthly()
        self.refresh_yearly()
        self.show_all_records()
        if hasattr(self, "graph_figure"):
            self.refresh_graphs()

    def refresh_graphs(self):
        if not MATPLOTLIB_AVAILABLE or self.graph_figure is None:
            QMessageBox.warning(
                self,
                "グラフ表示不可",
                "matplotlib が未インストールです。\n\npip install matplotlib\nを実行してください。"
            )
            return

        year = int(self.graph_year.currentText())
        month = int(self.graph_month.currentText())

        self.graph_figure.clear()

        ax1 = self.graph_figure.add_subplot(2, 1, 1)
        series = self.db.monthly_profit_series(year, self.salary_day)
        months = [m for m, _, _, _, _ in series]
        profits = [p for _, p, _, _, _ in series]

        ax1.plot(months, profits, marker="o")
        ax1.axhline(0, linewidth=1)
        ax1.set_title(f"{year}年 月次収益推移")
        ax1.set_xlabel("月")
        ax1.set_ylabel("収益")
        ax1.set_xticks(months)
        ax1.grid(True, alpha=0.3)

        ax2 = self.graph_figure.add_subplot(2, 1, 2)
        start, end = monthly_period(year, month, self.salary_day)
        rows = self.db.category_expense_between(start, end)

        labels = [str(r["category2"]) for r in rows]
        values = [int(r["total_expense"]) for r in rows]

        if values:
            ax2.pie(values, labels=labels, autopct="%1.1f%%", startangle=90)
            ax2.set_title(f"{year}年{month}月 カテゴリ別支出")
        else:
            ax2.text(0.5, 0.5, "対象データなし", ha="center", va="center")
            ax2.set_title(f"{year}年{month}月 カテゴリ別支出")

        self.graph_figure.tight_layout()
        self.graph_canvas.draw()

    def open_budget_config(self):
        dialog = BudgetConfigDialog(self.db, self)
        dialog.exec()
        self.refresh_master_combos()
        self.refresh_monthly()
        self.refresh_graphs()

    def open_recurring_config(self):
        dialog = RecurringConfigDialog(self.db, self)
        dialog.exec()
        self.refresh_master_combos()
        self.refresh_all()
        self.refresh_graphs()

    def open_master_config(self, master_type: str):
        dialog = MasterConfigDialog(self.db, master_type, self)
        dialog.exec()

        self.refresh_master_combos()

    def open_salary_day_config(self):
        dialog = SalaryDayConfigDialog(self.salary_day, self)

        if dialog.exec() != QDialog.Accepted:
            return

        new_day = dialog.salary_day()

        try:
            self.db.set_salary_day(new_day)
        except Exception as e:
            QMessageBox.critical(self, "設定エラー", str(e))
            return

        self.salary_day = new_day
        self.refresh_all()

        QMessageBox.information(
            self,
            "設定完了",
            f"給与支給日を {new_day} 日に設定しました。\n"
            "土曜は直前の金曜、日曜は翌月曜に補正して月間集計します。"
        )

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Excelファイルを選択",
            str(APP_DIR),
            "Excel Files (*.xlsx *.xlsm)"
        )

        if not path:
            return

        try:
            count = self.db.import_excel(path)
        except Exception as e:
            QMessageBox.critical(self, "取込エラー", str(e))
            return

        self.refresh_master_combos()
        self.refresh_all()

        QMessageBox.information(self, "取込完了", f"{count}件取り込みました。")


def main():
    app = QApplication(sys.argv)
    apply_app_icon(app)

    app.setStyleSheet("""
        QMainWindow {
            background:#f5f7fb;
        }

        QGroupBox {
            font-weight:bold;
            border:1px solid #d0d7de;
            border-radius:8px;
            margin-top:12px;
            background:white;
        }

        QGroupBox::title {
            subcontrol-origin:margin;
            left:12px;
            padding:0 4px;
        }

        QPushButton {
            padding:8px 14px;
            border-radius:6px;
            background:#0aa34f;
            color:white;
            font-weight:bold;
        }

        QPushButton:hover {
            background:#088d45;
        }

        QLineEdit,
        QComboBox,
        QDateEdit,
        QSpinBox {
            padding:6px;
            border:1px solid #c9d1d9;
            border-radius:4px;
            background:white;
        }

        QTableWidget {
            background:white;
            gridline-color:#d0d7de;
        }

        QHeaderView::section {
            background:#00a84f;
            color:white;
            font-weight:bold;
            padding:6px;
            border:1px solid #d0d7de;
        }
    """)

    w = MainWindow()
    w.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
